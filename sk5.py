import openpyxl

wb = openpyxl.load_workbook("ammu.xlsx")
ws = wb.active
# # ws["A1"] = "ammu"
# l = ["ammu","harshini","ashwin","gopika","haritha"]
# ws.append(l)
# ws.cell(row=3,column=2,value="anu")

# ws = wb.create_sheet("aadhu",1)
a = wb.sheetnames
ws = wb["aasha"]
# print(a)
ws['a1'] = 'Ammu'
ws.title = "aasha"
r = ws.rows
for i in r:
    for j in i:
        print(j.value)
    print()

ws["G4"] = "=sum(A4:F4)"
wb.save("ammu.xlsx")